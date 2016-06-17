#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import os, os.path
import json
import shutil
import openpyxl
from PIL import Image

PRIVACY_URL = "http://senspark.com/privacy-policy"
MARKETING_URL = "http://senspark.com"
SUPPORT_URL = "http://fb.com/teamsenspark"

IOS_LANGUAGES_CODES   = {       'en-US'     :   'english',
                                'en-AU'     :   'english',
                                'en-CA'     :   'english',
                                'en-GB'     :   'english',
                                'zh-Hans'   :   'chinese simplified',
                                'zh-Hant'   :   'chinese traditional',
                                'da'        :   'danish',
                                'de-DE'     :   'german',
                                'el'        :   'greek',
                                'fi'        :   'finnish',
                                'fr-CA'     :   'french',
                                'fr-FR'     :   'french',
                                'id'        :   'indonesian',
                                'it'        :   'italian',
                                'ja'        :   'japanese',
                                'ko'        :   'korean',
                                'ms'        :   'malay',
                                'nl-NL'     :   'dutch',
                                'no'        :   'norwegian',
                                'pt-BR'     :   'portuguese brazil',
                                'pt-PT'     :   'portuguese portugal',
                                'ru'        :   'russian',
                                'es-MX'     :   'spanish',
                                'es-ES'     :   'spanish',
                                'sv'        :   'swedish',
                                'th'        :   'thai',
                                'tr'        :   'turkish',
                                'vi'        :   'vietnamese'
}

ANDROID_LANGUAGES_CODES   = {   'en-US'     :   'english',
                                'zh-CN'     :   'chinese simplified',
                                'zh-TW'     :   'chinese traditional',
                                'ar'        :   'arabic',
                                'be'        :   'belarusian',
                                'bg'        :   'bulgarian',
                                'cs-CZ'     :   'czech',
                                'da-DK'     :   'danish',
                                'de-DE'     :   'german',
                                'el-GR'     :   'greek',
                                'es-ES'     :   'spanish',
                                'fi-FI'     :   'finnish',
                                'fil'       :   'filipino',
                                'fr-FR'     :   'french',
                                'hi-IN'     :   'hindi',
                                'hr'        :   'croatian',
                                'hu-HU'     :   'hungarian',
                                'id'        :   'indonesian',
                                'it-IT'     :   'italian',
                                'ja-JP'     :   'japanese',
                                'ko-KR'     :   'korean',
                                'lv'        :   'latvian',
                                'ms'        :   'malay',
                                'nl-NL'     :   'dutch',
                                'no-NO'     :   'norwegian',
                                'pl-PL'     :   'Polish',
                                'pt-BR'     :   'portuguese brazil',
                                'pt-PT'     :   'portuguese portugal',
                                'ro'        :   'romanian',
                                'ru-RU'     :   'russian',
                                'sr'        :   'serbian',
                                'sk'        :   'slovak',
                                'sl'        :   'slovenian',
                                'sv-SE'     :   'swedish',
                                'th'        :   'thai',
                                'tr-TR'     :   'turkish',
                                'uk'        :   'ukrainian',
                                'vi'        :   'vietnamese',
                                'af'        :   'afrikaans',
                                'am'        :   'amharic',
                                'hy-AM'     :   'armenian',
                                'az-AZ'     :   'azerbaijani',
                                'eu-ES'     :   'basque',
                                'bn-BD'     :   'bengali',
                                'my-MM'     :   'burmese',
                                'ca'        :   'catalan',
                                'et'        :   'estonian',
                                'gl-ES'     :   'galician',
                                'ka-GE'     :   'georgian',
                                'iw-IL'     :   'hebrew',
                                'is-IS'     :   'icelandic',
                                'kn-IN'     :   'kannada',
                                'km-KH'     :   'khmer',
                                'ky-KG'     :   'kyrgyz',
                                'lo-LA'     :   'lao',
                                'lt'        :   'lithuanian',
                                'mk-MK'     :   'macedonian',
                                'ml-IN'     :   'malayalam',
                                'mr-IN'     :   'marathi',
                                'mn-MN'     :   'mongolian',
                                'ne-NP'     :   'nepali',
                                'fa'        :   'persian',
                                'si-LK'     :   'Sinhala',
                                'sw'        :   'swahili',
                                'ta-IN'     :   'tamil',
                                'te-IN'     :   'telugu',
                                'zu'        :   'zulu',
}

context = {}

def checkParams(requiredParams, optionalParams, context):
    for i in range(1, len(sys.argv)):
        for j in range(0, len(requiredParams)):
            if "-" + requiredParams[j] == sys.argv[i]:
                context[requiredParams[j]] = sys.argv[i+1]
        for j in range(0, len(optionalParams)):
            if "-" + optionalParams[j] == sys.argv[i]:
                context[optionalParams[j]] = sys.argv[i+1]

    # print error log out required paramters are not ready
    raise_error = False;
    for i in range(0, len(requiredParams)):
        if context[requiredParams[i]] == "undefined":
            print "Invalid " + requiredParams[i] + " parameter"
            raise_error = True;

    if raise_error != False:
        dumpUsage()
        sys.exit()
        
def dumpUsage():
    print "usage:   ./populate (metadata|screenshots) -platform <platform> -prj-path <prj-path> <parameters>"
    print "\n"
    print "metadata command parameters"
    print "         -data-file-path <data-file-path>            xlsx data file path"
    print "         [-customized-metadata-path <customized-metadata-path>]"
    print "                                                     after populate metadata from xlsx data, will overwrite with this customized metadata"
    print "sample:  ./populate.py metadata -platform iOS -prj-path . -data-file-path ../src/data.xlsx -customized-metadata-path ../src/itunes/metadata"
    print "\n"
    print "screenshots command parameters"
    print "         -screenshots-path <screenshots-path>        screenshots path"
    print "         [-customized-screenshots-path <customized-screenshots-path>]"                                                                                                                    
    print "sample:  ./populate.py screenshots -platform android -prj-path . -screenshots-path ../src/screenshots -customized-screenshots-path"

def genFile(path, content):
    print path
    file = open(path, 'w')
    file.write(content)
    file.close()

def checkAndGenFile(path, content, max):
    if len(content) > max:
        print "%s is too long and greater than %d characters: %d" % (path, max, len(content))
        #sys.exit()
    genFile(path, content.encode('utf-8'))

# populate metadata
def populateMetadata():
    context = {
        'data-file-path'            : 'undefined',
        'prj-path'                  : 'undefined',
        'customized-metadata-path'  : 'undefined',
        'platform'                  : 'undefined',
    }

    checkParams(['data-file-path', 'prj-path', 'platform'], ['customized-metadata-path'], context)
    metadataPath = context['prj-path'] + '/metadata/'
    cmd = 'rm -r ' + metadataPath
    print cmd
    os.system(cmd);

    if context['platform'].lower()=='ios':
        for key, value in IOS_LANGUAGES_CODES.iteritems():
            wb = openpyxl.load_workbook(context['data-file-path'])
            sheet = wb.get_sheet_by_name('Sheet1')
            
            #prepare languages folders
            if not os.path.exists(metadataPath + key):
                os.makedirs(metadataPath + key)
            
            #write privacy_url, marketing_url, support_url files
            genFile(metadataPath + key + '/privacy_url.txt', PRIVACY_URL)
            genFile(metadataPath + key + '/marketing_url.txt', MARKETING_URL)
            genFile(metadataPath + key + '/support_url.txt', SUPPORT_URL)
            
            #check loading excel file
            #print sheet.cell(row = 1, column = 2).value + " "
            
            #write names, keywords, release notes, description
            foundLang = 'false'
            for i in range(2, 29):
                if str(sheet.cell(row=1, column=i).value).lower() == value.lower():
                    foundLang = 'true'
                    checkAndGenFile(metadataPath + key + '/name.txt', sheet.cell(row=2, column=i).value, 255)
                    checkAndGenFile(metadataPath + key + '/keywords.txt', sheet.cell(row=4, column=i).value, 100)
                    checkAndGenFile(metadataPath + key + '/release_notes.txt', sheet.cell(row=5, column=i).value, 4000)
                    checkAndGenFile(metadataPath + key + '/description.txt', sheet.cell(row=7, column=i).value, 4000)
            if foundLang == 'false':
                print "Not found data for language %s %" % (value, key)
                sys.exit()
                
        if (context['customized-metadata-path']!="undefined"):
            cmd = 'cp -r ' + context['customized-metadata-path'] + '/ ' + metadataPath
            print cmd
            os.system(cmd)
                     
    elif context['platform'].lower()=='android':
        for key, value in ANDROID_LANGUAGES_CODES.iteritems():
            wb = openpyxl.load_workbook(context['data-file-path'])
            sheet = wb.get_sheet_by_name('Sheet1')
            
            #prepare languages folders
            if not os.path.exists(metadataPath + key):
                os.makedirs(metadataPath + key)
                
            #write privacy_url, marketing_url, support_url files
            if len(sheet.cell(row=8, column=2).value) > 4:
                genFile(metadataPath + key + '/video.txt', sheet.cell(row=8, column=2).value)
            
            #check loading excel file
            #print sheet.cell(row = 1, column = 2).value + " "
            
            #write title, short description, full description
            foundLang = 'false'
            for i in range(2, 69):
                if str(sheet.cell(row=1, column=i).value).lower() == value.lower(): # find the language in excel data
                    foundLang = 'true'
                     
                    if len(sheet.cell(row=2, column=i).value)<=30:
                        checkAndGenFile(metadataPath + key + '/title.txt', sheet.cell(row=2, column=i).value, 30)
                    else:
                        checkAndGenFile(metadataPath + key + '/title.txt', sheet.cell(row=3, column=i).value, 30)
                    checkAndGenFile(metadataPath + key + '/short_description.txt', sheet.cell(row=6, column=i).value, 80)
                    checkAndGenFile(metadataPath + key + '/full_description.txt', sheet.cell(row=7, column=i).value, 4000)
                    if sheet.cell(row=9, column=2).value!="": # write changelog for android
                        if not os.path.exists(metadataPath + key + "/changelogs"):
                            os.makedirs(metadataPath + key + "/changelogs")
                        else:
                            'to do or not'
                            #cmd = 'rm -r ' + metadataPath + key + "/changelogs/"
                            #print cmd
                            #os.system(cmd);
                            
                        verCodes = str(sheet.cell(row=9, column=2).value).split()
                        for verCode in verCodes:
                            checkAndGenFile(metadataPath + key + '/changelogs/' + verCode + '.txt', sheet.cell(row=5, column=i).value, 500)
            if foundLang == 'false':
                print "Not found data for language %s %" % (value, key)
                sys.exit()
                    
        if (context['customized-metadata-path']!="undefined"):
            cmd = 'cp -r ' + context['customized-metadata-path'] + '/ ' + metadataPath
            print cmd
            os.system(cmd)

# populate screenshots, TODO: code for android
def populateScreenshots():
    context = {
        'screenshots-path'  : 'undefined',
        'prj-path'          : 'undefined',
        'customized-screenshots-path'  : 'undefined',
        'platform'                  : 'undefined',
    }
    
    checkParams(['screenshots-path', 'prj-path'], ['customized-screenshots-path'], context)
    for root, dirs, files in os.walk(context['screenshots-path'], topdown=False):
        for name in files:
            if name.endswith(".png"):
                source_path = os.path.join(root, name)
                im = Image.open(source_path)
                target_path = source_path[:-3] + 'jpg'
                if not os.path.exists(target_path):
                    im.save(target_path, "JPEG")
                    print "converted %s to jpg" % source_path

    count = 0
    
    screenshotsPath = context['prj-path'] + '/screenshots/'
    cmd = 'rm -r ' + screenshotsPath
    print cmd
    os.system(cmd)
    
    for key, value in IOS_LANGUAGES_CODES.iteritems():
        #prepare languages folders
        if not os.path.exists(screenshotsPath + key):
            os.makedirs(screenshotsPath + key)
        
        for root, dirs, files in os.walk(context['screenshots-path'], topdown=False):
            for name in files:
                if name.endswith(".jpg"):
                    count = count + 1
                    source_path = os.path.join(root, name)
                    target_path = screenshotsPath + key + '/' + name
                    shutil.copyfile(source_path, target_path)
                    print "Copy file: %s > %s" % (source_path, target_path)
    if (context['customized-screenshots-path']!="undefined"):
        cmd = 'cp -r ' + context['customized-screenshots-path'] + '/ ' + screenshotsPath
        print cmd
        os.system(cmd)
            
def selectCommand():
    #print sys.argv[0]
    if len(sys.argv)<2:
        dumpUsage()
        sys.exit()
    if "metadata" == sys.argv[1]:
        populateMetadata()
    elif "screenshots" == sys.argv[1]:
        populateScreenshots()

selectCommand()